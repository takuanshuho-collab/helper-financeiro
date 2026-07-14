"""
Geração da planilha .xlsx (diagnóstico + simulações).

Filosofia: a planilha não é uma foto morta. As entradas ficam em células
editáveis e os indicadores derivados são FÓRMULAS do Excel — então o usuário
pode mudar um saldo ou uma taxa e ver tudo recalcular sozinho.
"""
from __future__ import annotations

from collections.abc import Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.estrategias import comparar_estrategias, oportunidades_portabilidade
from core.models import PerfilFinanceiro
from core.rubricas import ROTULO_CAMPO, ROTULO_CATEGORIA, Rubrica

# --- Paleta e estilos ---
AZUL = "1F4E79"
AZUL_CLARO = "DDEBF7"
CINZA = "F2F2F2"
FONTE = "Calibri"

MOEDA = '"R$" #,##0.00'
PCT = "0.00%"

_titulo = Font(name=FONTE, bold=True, color="FFFFFF", size=12)
_cabecalho = Font(name=FONTE, bold=True, color="FFFFFF")
_entrada = Font(name=FONTE, color="0000FF")     # azul = valor que o usuário edita
_normal = Font(name=FONTE)
_negrito = Font(name=FONTE, bold=True)
_fill_titulo = PatternFill("solid", fgColor=AZUL)
_fill_cab = PatternFill("solid", fgColor=AZUL)
_fill_entrada = PatternFill("solid", fgColor="FFF2CC")  # amarelo suave = editável
_borda = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def _mesclar_titulo(ws, celula, texto, ate_coluna):
    ws[celula] = texto
    ws[celula].font = _titulo
    ws[celula].fill = _fill_titulo
    linha = celula[1:]
    ws.merge_cells(f"{celula}:{ate_coluna}{linha}")


def _aba_dividas(wb, perfil: PerfilFinanceiro):
    ws = wb.create_sheet("Dívidas")
    _mesclar_titulo(ws, "A1", "DÍVIDAS", "I")

    cabecalhos = ["Credor", "Tipo", "Saldo devedor", "Taxa a.m.", "Taxa a.a.",
                  "Parcela", "Parc. restantes", "Custo total restante", "Juros restantes"]
    ws.append([])
    ws.append(cabecalhos)
    for cell in ws[3]:
        cell.font = _cabecalho
        cell.fill = _fill_cab
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    primeira = 4
    for i, d in enumerate(perfil.dividas):
        linha = primeira + i
        ws.cell(linha, 1, d.credor).font = _normal
        ws.cell(linha, 2, d.tipo).font = _normal
        ws.cell(linha, 3, round(d.saldo_devedor, 2)).font = _entrada     # entrada
        ws.cell(linha, 4, round(d.taxa_mensal, 6)).font = _entrada       # entrada
        # Taxa anual = (1+mensal)^12 - 1  (fórmula, referencia a coluna D)
        ws.cell(linha, 5, f"=(1+D{linha})^12-1").font = _normal
        ws.cell(linha, 6, round(d.parcela, 2)).font = _entrada           # entrada
        ws.cell(linha, 7, d.parcelas_restantes).font = _entrada          # entrada
        ws.cell(linha, 8, f"=F{linha}*G{linha}").font = _normal          # custo total
        ws.cell(linha, 9, f"=MAX(H{linha}-C{linha},0)").font = _normal   # juros restantes

        for col in (3, 6, 8, 9):
            ws.cell(linha, col).number_format = MOEDA
        for col in (4, 5):
            ws.cell(linha, col).number_format = PCT

    ultima = primeira + len(perfil.dividas) - 1
    # Linha de total
    tot = ultima + 1
    ws.cell(tot, 2, "TOTAL").font = _negrito
    for col, letra in ((3, "C"), (6, "F"), (8, "H"), (9, "I")):
        c = ws.cell(tot, col, f"=SUM({letra}{primeira}:{letra}{ultima})")
        c.font = _negrito
        c.number_format = MOEDA

    larguras = [22, 30, 16, 10, 10, 14, 14, 20, 18]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Gráfico de barras: saldo devedor por credor
    if perfil.dividas:
        chart = BarChart()
        chart.title = "Saldo devedor por dívida"
        chart.type = "bar"
        dados = Reference(ws, min_col=3, min_row=3, max_row=ultima)
        cats = Reference(ws, min_col=1, min_row=primeira, max_row=ultima)
        chart.add_data(dados, titles_from_data=True)
        chart.set_categories(cats)
        chart.legend = None
        chart.height = 6
        chart.width = 14
        ws.add_chart(chart, f"A{tot + 3}")

    return ws, primeira, ultima


def _aba_diagnostico(wb, perfil: PerfilFinanceiro, primeira: int, ultima: int):
    ws = wb.active
    ws.title = "Diagnóstico"
    _mesclar_titulo(ws, "A1", "DIAGNÓSTICO FINANCEIRO", "C")

    linhas = [
        ("Renda líquida mensal", round(perfil.renda_liquida, 2), True, MOEDA),
        ("Despesas fixas", round(perfil.despesas_fixas, 2), True, MOEDA),
        ("Despesas variáveis", round(perfil.despesas_variaveis, 2), True, MOEDA),
        ("Reserva de emergência", round(perfil.reserva_emergencia, 2), True, MOEDA),
        ("Saldo de FGTS", round(perfil.saldo_fgts, 2), True, MOEDA),
    ]
    r = 3
    for rotulo, valor, entrada, fmt in linhas:
        ws.cell(r, 1, rotulo).font = _negrito
        c = ws.cell(r, 2, valor)
        c.font = _entrada if entrada else _normal
        if entrada:
            c.fill = _fill_entrada
        c.number_format = fmt
        r += 1

    # Derivados (fórmulas). Total de parcelas vem da aba Dívidas.
    faixa_parcelas = f"'Dívidas'!F{primeira}:F{ultima}"
    faixa_saldos = f"'Dívidas'!C{primeira}:C{ultima}"
    derivados = [
        ("Total de parcelas/mês", f"=SUM({faixa_parcelas})", MOEDA),
        ("Despesas totais", "=B4+B5", MOEDA),
        ("Fluxo de caixa (sobra/mês)", "=B3-B9-B8", MOEDA),
        ("Saldo devedor total", f"=SUM({faixa_saldos})", MOEDA),
        ("Comprometimento de renda", "=B8/B3", PCT),
    ]
    for rotulo, formula, fmt in derivados:
        ws.cell(r, 1, rotulo).font = _negrito
        c = ws.cell(r, 2, formula)
        c.font = _normal
        c.number_format = fmt
        r += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    nota = ws.cell(r + 1, 1,
                   "Células em amarelo são entradas: altere-as e os indicadores "
                   "recalculam automaticamente ao abrir no Excel.")
    nota.font = Font(name=FONTE, italic=True, size=9, color="808080")
    return ws


def _aba_estrategias(wb, perfil: PerfilFinanceiro, extra_mensal: float,
                     taxa_alvo: float):
    ws = wb.create_sheet("Estratégias")
    _mesclar_titulo(ws, "A1", "ESTRATÉGIAS DE QUITAÇÃO", "E")

    comp = comparar_estrategias(perfil, extra_mensal)
    ws.append([])
    ws.cell(3, 1, "Pagamento extra considerado por mês:").font = _negrito
    c = ws.cell(3, 2, round(extra_mensal, 2))
    c.number_format = MOEDA
    c.font = _entrada

    ws.append([])
    ws.append(["Método", "Meses até quitar", "Total de juros pagos", "Ordem de ataque"])
    for cell in ws[5]:
        cell.font = _cabecalho
        cell.fill = _fill_cab

    def _linha_metodo(nome, res):
        meses = res["meses"] if res["quitavel"] else "não quita c/ este valor"
        ordem = " → ".join(res["ordem"])
        ws.append([nome, meses, res["juros_pagos"], ordem])
        ws.cell(ws.max_row, 3).number_format = MOEDA

    _linha_metodo("Avalanche (maior juro primeiro)", comp["avalanche"])
    _linha_metodo("Bola de neve (menor saldo primeiro)", comp["bola_de_neve"])

    # Portabilidade
    ws.append([])
    ws.append([])
    lin_port = ws.max_row + 1
    ws.cell(lin_port, 1,
            f"Portabilidade — economia se migrar para {taxa_alvo*100:.2f}% a.m.").font = _negrito
    ws.append(["Credor", "Tipo", "Parcela atual", "Parcela nova", "Economia total"])
    for cell in ws[ws.max_row]:
        cell.font = _cabecalho
        cell.fill = _fill_cab
    for o in oportunidades_portabilidade(perfil, taxa_alvo):
        ws.append([o["credor"], o["tipo"], o["parcela_atual"],
                   o["parcela_nova"], o["economia_total"]])
        for col in (3, 4, 5):
            ws.cell(ws.max_row, col).number_format = MOEDA

    for i, w in enumerate([34, 30, 16, 16, 16], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    return ws


def _aba_orcamento(wb, rubricas: Sequence[Rubrica]):
    """Aba "Orçamento detalhado": as rubricas do usuário (ADR-0012).

    Segue a filosofia da planilha viva: o VALOR de cada rubrica é entrada
    editável e o subtotal por campo é uma fórmula =SUM — alterar uma conta no
    Excel recalcula o subtotal na hora.
    """
    ws = wb.create_sheet("Orçamento detalhado")
    _mesclar_titulo(ws, "A1", "ORÇAMENTO DETALHADO (RUBRICAS)", "C")

    ws.append([])
    ws.append(["Campo do orçamento", "Rubrica", "Valor"])
    for cell in ws[3]:
        cell.font = _cabecalho
        cell.fill = _fill_cab

    # Agrupa preservando a ordem canônica dos campos (a mesma da GUI).
    linha = 4
    for categoria, campos in ROTULO_CAMPO.items():
        for campo, rotulo in campos.items():
            do_campo = [r for r in rubricas
                        if r.categoria == categoria and r.campo_pai == campo]
            if not do_campo:
                continue
            primeira = linha
            for r in do_campo:
                ws.cell(linha, 1,
                        f"{ROTULO_CATEGORIA[categoria]} · {rotulo}").font = _normal
                ws.cell(linha, 2, r.nome).font = _normal
                c = ws.cell(linha, 3, round(r.valor, 2))
                c.font = _entrada
                c.fill = _fill_entrada
                c.number_format = MOEDA
                linha += 1
            sub = ws.cell(linha, 2, f"Subtotal — {rotulo}")
            sub.font = _negrito
            c = ws.cell(linha, 3, f"=SUM(C{primeira}:C{linha - 1})")
            c.font = _negrito
            c.number_format = MOEDA
            linha += 2

    nota = ws.cell(linha, 1,
                   "Este campo aparece com o subtotal na aba Diagnóstico; no "
                   "app, o campo detalhado vale a soma das rubricas.")
    nota.font = Font(name=FONTE, italic=True, size=9, color="808080")
    for letra, largura in (("A", 40), ("B", 32), ("C", 16)):
        ws.column_dimensions[letra].width = largura
    return ws


def _aba_evolucao_cabecalho(ws, meses: list[str], ultima_col: int) -> int:
    """Escreve a linha de cabeçalho (campo + competências) e retorna sua linha."""
    linha_cab = 3
    ws.cell(linha_cab, 1, "Campo do orçamento")
    for j, mes in enumerate(meses, start=2):
        ws.cell(linha_cab, j, mes)
    for cell in ws[linha_cab]:
        cell.font = _cabecalho
        cell.fill = _fill_cab
        cell.alignment = Alignment(horizontal="center")
    return linha_cab


def _aba_evolucao_secoes(ws, secoes, linha_cab: int,
                          ultima_col: int) -> tuple[int, list[tuple[str, int]]]:
    """Escreve as seções (campos editáveis + fórmula =SUM de total por coluna).

    Retorna a próxima linha livre e a lista (rótulo, linha do total) de cada
    seção — insumo do bloco-resumo que alimenta o gráfico.
    """
    linha = linha_cab + 1
    totais_por_secao: list[tuple[str, int]] = []  # (rótulo, linha do total)
    for secao in secoes:
        if not secao["campos"]:
            continue  # seção zerada em todo o período: fora (ruído)
        ws.cell(linha, 1, secao["rotulo"]).font = _negrito
        linha += 1
        primeira = linha
        for campo in secao["campos"]:
            ws.cell(linha, 1, campo["rotulo"]).font = _normal
            for j, valor in enumerate(campo["valores"], start=2):
                c = ws.cell(linha, j, round(valor, 2))
                c.font = _entrada
                c.fill = _fill_entrada
                c.number_format = MOEDA
            linha += 1
        ws.cell(linha, 1, f"Total — {secao['rotulo']}").font = _negrito
        for j in range(2, ultima_col + 1):
            letra = get_column_letter(j)
            c = ws.cell(linha, j, f"=SUM({letra}{primeira}:{letra}{linha - 1})")
            c.font = _negrito
            c.number_format = MOEDA
        totais_por_secao.append((secao["rotulo"], linha))
        linha += 2
    return linha, totais_por_secao


def _aba_evolucao_resumo(ws, totais_por_secao: list[tuple[str, int]], linha: int,
                          ultima_col: int) -> tuple[int, int, int]:
    """Bloco-resumo (referencia os totais) — é a fonte do gráfico de linhas.

    Retorna (primeira linha do resumo, última linha do resumo, próxima linha livre).
    """
    resumo_primeira = linha
    for rotulo, linha_total in totais_por_secao:
        ws.cell(linha, 1, rotulo).font = _normal
        for j in range(2, ultima_col + 1):
            letra = get_column_letter(j)
            c = ws.cell(linha, j, f"={letra}{linha_total}")
            c.number_format = MOEDA
        linha += 1
    resumo_ultima = linha - 1
    return resumo_primeira, resumo_ultima, linha


def _aba_evolucao_grafico(ws, resumo_primeira: int, resumo_ultima: int, linha_cab: int,
                           ultima_col: int, linha: int) -> None:
    """Monta o gráfico de linhas nativo a partir do bloco-resumo."""
    chart = LineChart()
    chart.title = "Evolução por seção"
    dados = Reference(ws, min_col=1, max_col=ultima_col,
                      min_row=resumo_primeira, max_row=resumo_ultima)
    chart.add_data(dados, titles_from_data=True, from_rows=True)
    cats = Reference(ws, min_col=2, max_col=ultima_col, min_row=linha_cab)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 18
    ws.add_chart(chart, f"A{linha + 2}")


def _aba_evolucao(wb, evolucao: dict):
    """Aba "Evolução mensal": campos × competências arquivadas (ADR-0014).

    Os valores dos campos são entradas editáveis e o total de cada seção é
    fórmula =SUM por coluna (planilha viva). Um bloco-resumo referencia os
    totais e alimenta o gráfico de linhas nativo do Excel.
    """
    meses: list[str] = evolucao["meses"]
    ultima_col = 1 + len(meses)

    ws = wb.create_sheet("Evolução mensal")
    _mesclar_titulo(ws, "A1", "EVOLUÇÃO MENSAL DO ORÇAMENTO",
                    get_column_letter(max(ultima_col, 2)))

    linha_cab = _aba_evolucao_cabecalho(ws, meses, ultima_col)
    linha, totais_por_secao = _aba_evolucao_secoes(ws, evolucao["secoes"], linha_cab, ultima_col)
    resumo_primeira, resumo_ultima, linha = _aba_evolucao_resumo(
        ws, totais_por_secao, linha, ultima_col)
    _aba_evolucao_grafico(ws, resumo_primeira, resumo_ultima, linha_cab, ultima_col, linha)

    ws.column_dimensions["A"].width = 30
    for j in range(2, ultima_col + 1):
        ws.column_dimensions[get_column_letter(j)].width = 14
    return ws


def gerar_planilha(perfil: PerfilFinanceiro, caminho_saida: str,
                   extra_mensal: float = 0.0, taxa_alvo_mensal: float = 0.018,
                   rubricas: Sequence[Rubrica] | None = None,
                   evolucao: dict | None = None) -> str:
    """Monta e salva a planilha completa. Retorna o caminho salvo.

    `evolucao` é a saída de `core.rubricas.serie_evolucao` (competências
    arquivadas); a aba "Evolução mensal" só existe quando há histórico.
    """
    wb = Workbook()
    _, primeira, ultima = _aba_dividas(wb, perfil)
    _aba_diagnostico(wb, perfil, primeira, ultima)
    _aba_estrategias(wb, perfil, extra_mensal, taxa_alvo_mensal)
    if rubricas:  # aba só existe quando o usuário detalhou o orçamento
        _aba_orcamento(wb, rubricas)
    if (evolucao and evolucao.get("meses")
            and any(s["campos"] for s in evolucao["secoes"])):
        _aba_evolucao(wb, evolucao)
    # Ordena as abas: Diagnóstico primeiro
    wb.move_sheet("Diagnóstico", -(wb.sheetnames.index("Diagnóstico")))
    wb.save(caminho_saida)
    return caminho_saida
