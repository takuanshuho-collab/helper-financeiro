"""
Testes de contrato do sidecar (REQ-NF-005 / REQ-SEC-004).

Usam o `TestClient` do FastAPI (sem rede real). Cobrem a autenticação por
token, a validação de entrada e o roundtrip determinístico core <-> JSON,
incluindo os casos de borda (sem dívidas, reserva sem despesas, ordenação).
"""
import base64
import os
import time

import pytest
from fastapi.testclient import TestClient

from agent.config import ConfigAgente
from agent.provider import FakeProvider
from contracts import (
    AnaliseAgente,
    ClassificacaoExtrato,
    FatosFinanceiros,
    ItemClassificado,
)
from sidecar.app import (
    app,
    contexto_analise,
    contexto_classificacao,
    contexto_extracao,
    contexto_ocr,
    exigir_cofre,
    sessao_dependencia,
)
from sidecar.auth import Cofre
from sidecar.persistencia import Repositorio
from sidecar.security import VAR_TOKEN
from sidecar.sessao import SessaoCofre
from tests.test_classificacao import FakeClassificador
from tests.test_extracao import CFG_TESTE, DOC_CONTRATO, FakeExtrator
from tests.test_ocr import MotorFalso, _linha

TOKEN = "token-de-teste"
CABECALHO = {"X-HF-Token": TOKEN}
cliente = TestClient(app)

# PDF fantasma: o texto real é injetado via monkeypatch de extrair_texto_pdf_bytes.
PDF_B64 = base64.b64encode(b"%PDF-1.4 fake").decode()


def setup_module(_module):
    os.environ[VAR_TOKEN] = TOKEN


# --- Sessão do cofre (T-1603): janela de onboarding por padrão ---------------
# Todo teste deste módulo, por padrão, roda "sem cofre cadastrado" — o mesmo
# comportamento de antes do T-1603 (endpoints de negócio abertos). Os testes
# do cofre em si (tests/test_sessao.py) sobrescrevem `sessao_dependencia` com
# uma sessão própria (cadastrada/bloqueada/desbloqueada conforme o cenário).
@pytest.fixture(autouse=True)
def _sessao_sem_cofre(tmp_path):
    sess = SessaoCofre(
        cofre=Cofre(tmp_path / "_autouse_auth.json"),
        caminho_db=tmp_path / "_autouse_dados.db",
    )
    app.dependency_overrides[sessao_dependencia] = lambda: sess
    yield sess
    app.dependency_overrides.pop(sessao_dependencia, None)
    sess.fechar()


# --- Autenticação (REQ-SEC-004) ----------------------------------------------


def test_health_dispensa_token():
    resposta = cliente.get("/health")
    assert resposta.status_code == 200
    assert resposta.json()["status"] == "ok"


def test_diagnostico_sem_token_401():
    resposta = cliente.post("/diagnostico", json={})
    assert resposta.status_code == 401


def test_diagnostico_token_invalido_401():
    resposta = cliente.post(
        "/diagnostico", json={}, headers={"X-HF-Token": "errado"}
    )
    assert resposta.status_code == 401


# --- Encerramento gracioso (C-11) --------------------------------------------
# O Electron pede `POST /encerrar` antes do kill duro para o lifespan rodar
# (fecha SQLCipher, derruba o llama-server). O launcher injeta o `uvicorn.Server`
# em `app.state.servidor`; aqui usamos um duplo com `should_exit`.
class _ServidorFalso:
    def __init__(self) -> None:
        self.should_exit = False


def test_encerrar_sem_token_401():
    """Encerramento exige o token do processo (REQ-SEC-004): sem ele, 401 —
    nada de derrubar o serviço a partir de um request sem credencial."""
    resposta = cliente.post("/encerrar")
    assert resposta.status_code == 401


def test_encerrar_sinaliza_shutdown_do_servidor():
    """Com token, o endpoint seta `should_exit` no servidor injetado — é o que
    tira o uvicorn do loop e dispara o shutdown do lifespan."""
    servidor = _ServidorFalso()
    app.state.servidor = servidor
    try:
        resposta = cliente.post("/encerrar", headers=CABECALHO)
        assert resposta.status_code == 200
        assert resposta.json() == {"ok": True}
        assert servidor.should_exit is True
    finally:
        del app.state.servidor


def test_encerrar_sem_servidor_injetado_nao_quebra():
    """Sob TestClient não há `app.state.servidor` (nem loop uvicorn para
    encerrar): o endpoint apenas responde `ok`, sem AttributeError."""
    try:
        del app.state.servidor
    except (AttributeError, KeyError):
        pass
    resposta = cliente.post("/encerrar", headers=CABECALHO)
    assert resposta.status_code == 200
    assert resposta.json() == {"ok": True}


# --- Validação de entrada (REQ-NF-005) ---------------------------------------


def test_payload_invalido_422():
    # Dívida sem os campos obrigatórios `credor`/`tipo`.
    payload = {"dividas": [{"saldo_devedor": 100.0}]}
    resposta = cliente.post("/diagnostico", json=payload, headers=CABECALHO)
    assert resposta.status_code == 422


def test_payload_invalido_detail_e_string_422():
    """C-07 (alto), independente de C-01: QUALQUER 422 de validação
    automática do Pydantic — mesmo o de campo obrigatório ausente, que já
    existia antes deste ciclo — precisa devolver `detail` como STRING. O
    handler padrão do FastAPI (não sobrescrito) devolve uma LISTA de
    `{loc, msg}`; sem o `exception_handler(RequestValidationError)`, o
    `client.ts` (que tipa `detail: string`) recebia a lista crua."""
    payload = {"dividas": [{"saldo_devedor": 100.0}]}  # sem credor/tipo
    resposta = cliente.post("/diagnostico", json=payload, headers=CABECALHO)
    assert resposta.status_code == 422
    detail = resposta.json()["detail"]
    assert isinstance(detail, str)


def test_valor_negativo_barrado_422():
    """C-01 (crítico): saldo_devedor negativo é barrado na fronteira, ANTES
    do core rodar — antes da correção (`Field(ge=0)`), este payload voltava
    200 com um número financeiro incorreto, e pior:
    `Divida.juros_restantes = max(custo_total_restante - saldo_devedor, 0.0)`
    MASCARAVA o erro como 0.0 em vez de sinalizá-lo (viola H1)."""
    payload = {"dividas": [
        {"credor": "Banco X", "tipo": "Empréstimo",
         "saldo_devedor": -100.0, "taxa_mensal": 0.02, "parcela": 50.0,
         "parcelas_restantes": 10},
    ]}
    resposta = cliente.post("/diagnostico", json=payload, headers=CABECALHO)
    assert resposta.status_code == 422
    # Nenhum ranking/`juros_restantes` chega a ser calculado: a entrada nunca
    # alcança o core (a barreira é o próprio Pydantic, não uma regra de
    # negócio) — o corpo do 422 não tem nenhum resquício do resumo.
    corpo = resposta.json()
    assert "ranking" not in corpo
    assert "divida_mais_cara" not in corpo


def test_valor_negativo_detail_e_string_422():
    """C-07 (alto): o `detail` de um 422 de validação automática do Pydantic
    é STRING legível, não a lista bruta `[{"loc":...,"msg":...}]` — sem o
    `exception_handler(RequestValidationError)`, o `client.ts` (que tipa
    `detail: string`) recebia a lista crua e o `Error()` a coagia para
    "[object Object]" na tela. Acoplado a C-01: o `Field(ge=0)` multiplica a
    frequência desse 422."""
    payload = {"dividas": [
        {"credor": "Banco X", "tipo": "Empréstimo", "saldo_devedor": -100.0},
    ]}
    resposta = cliente.post("/diagnostico", json=payload, headers=CABECALHO)
    assert resposta.status_code == 422
    detail = resposta.json()["detail"]
    assert isinstance(detail, str)
    assert "saldo_devedor" in detail


# --- Roundtrip determinístico (REQ-NF-005) -----------------------------------


def test_diagnostico_roundtrip():
    payload = {
        "renda": {"salario_liquido": 5000.0},
        "fixas": {"moradia": 1500.0, "contas_casa": 500.0},
        "variaveis": {"mercado": 800.0},
        "reserva_emergencia": 6000.0,
        "dividas": [
            {
                "credor": "Banco X",
                "tipo": "Cartão de crédito",
                "saldo_devedor": 10000.0,
                "taxa_mensal": 0.09,
                "parcela": 1200.0,
                "parcelas_restantes": 12,
            }
        ],
    }
    resposta = cliente.post("/diagnostico", json=payload, headers=CABECALHO)
    assert resposta.status_code == 200
    dados = resposta.json()

    # Roll-up determinístico no core: despesas = 1500 + 500 + 800 = 2800.
    assert dados["despesas_totais"] == 2800.0
    assert dados["despesas_fixas"] == 2000.0
    assert dados["despesas_variaveis"] == 800.0
    assert dados["total_parcelas"] == 1200.0
    # Comprometimento = 1200 / 5000 = 0,24 → Saudável.
    assert dados["classificacao"] == "Saudável"
    # Cobertura da reserva = 6000 / 2800 ≈ 2,14 meses.
    assert 2.0 < dados["meses_reserva"] < 2.3

    # Campos derivados da dívida foram serializados.
    mais_cara = dados["divida_mais_cara"]
    assert mais_cara["credor"] == "Banco X"
    assert mais_cara["taxa_anual"] > 0
    assert mais_cara["custo_total_restante"] == 1200.0 * 12

    # Estatísticas ponderadas da tela Dívidas (T-804), calculadas no core.
    assert dados["custo_total_ate_quitar"] == 1200.0 * 12
    # Uma única dívida: a média ponderada é a própria taxa.
    assert abs(dados["taxa_media_ponderada"] - 0.09) < 1e-9


def test_diagnostico_sem_dividas():
    payload = {
        "renda": {"salario_liquido": 4000.0},
        "fixas": {"moradia": 1000.0},
    }
    resposta = cliente.post("/diagnostico", json=payload, headers=CABECALHO)
    assert resposta.status_code == 200
    dados = resposta.json()

    assert dados["divida_mais_cara"] is None
    assert dados["ranking"] == []
    assert dados["saldo_devedor_total"] == 0.0
    assert dados["total_parcelas"] == 0.0
    assert dados["classificacao"] == "Saudável"


def test_meses_reserva_nulo_sem_despesas():
    # Sem despesas informadas, a cobertura em meses não tem significado.
    payload = {"renda": {"salario_liquido": 3000.0}, "reserva_emergencia": 5000.0}
    resposta = cliente.post("/diagnostico", json=payload, headers=CABECALHO)
    assert resposta.status_code == 200
    assert resposta.json()["meses_reserva"] is None


def test_ranking_ordena_por_taxa():
    payload = {
        "renda": {"salario_liquido": 9000.0},
        "dividas": [
            {
                "credor": "Consignado",
                "tipo": "Consignado",
                "saldo_devedor": 6000.0,
                "taxa_mensal": 0.018,
                "parcela": 350.0,
                "parcelas_restantes": 20,
            },
            {
                "credor": "Cartão",
                "tipo": "Cartão de crédito",
                "saldo_devedor": 8000.0,
                "taxa_mensal": 0.12,
                "parcela": 900.0,
                "parcelas_restantes": 12,
            },
        ],
    }
    resposta = cliente.post("/diagnostico", json=payload, headers=CABECALHO)
    assert resposta.status_code == 200
    dados = resposta.json()

    # Avalanche: a mais cara (0,12) vem primeiro.
    assert dados["divida_mais_cara"]["credor"] == "Cartão"
    assert [d["credor"] for d in dados["ranking"]] == ["Cartão", "Consignado"]


# --- Estratégias (avalanche vs. bola de neve) --------------------------------


def test_estrategias_sem_token_401():
    resposta = cliente.post("/estrategias", json={"perfil": {}})
    assert resposta.status_code == 401


def test_estrategias_compara_metodos():
    payload = {
        "perfil": {
            "renda": {"salario_liquido": 5000.0},
            "dividas": [
                {
                    "credor": "Cartão",
                    "tipo": "Cartão de crédito",
                    "saldo_devedor": 3000.0,
                    "taxa_mensal": 0.12,
                    "parcela": 600.0,
                    "parcelas_restantes": 8,
                },
                {
                    "credor": "Consignado",
                    "tipo": "Consignado",
                    "saldo_devedor": 5000.0,
                    "taxa_mensal": 0.018,
                    "parcela": 300.0,
                    "parcelas_restantes": 20,
                },
            ],
        },
        "extra": 200.0,
    }
    resposta = cliente.post("/estrategias", json=payload, headers=CABECALHO)
    assert resposta.status_code == 200
    dados = resposta.json()

    assert set(dados) == {"avalanche", "bola_de_neve"}
    # Avalanche ataca a mais cara primeiro (Cartão, 12% a.m.).
    assert dados["avalanche"]["ordem"][0] == "Cartão"
    assert dados["avalanche"]["quitavel"] is True
    assert dados["avalanche"]["meses"] is not None


# --- Contrato PDF: extração local + interrupt→resume (REQ-F-014) --------------


def test_contrato_sem_token_401():
    resposta = cliente.post("/contrato/extrair", json={"pdf_base64": PDF_B64})
    assert resposta.status_code == 401


def test_contrato_base64_invalido_422():
    cliente.app.dependency_overrides[contexto_extracao] = lambda: (CFG_TESTE, FakeExtrator())
    try:
        resposta = cliente.post(
            "/contrato/extrair", json={"pdf_base64": "nao-e-base64!!"}, headers=CABECALHO
        )
        assert resposta.status_code == 422
    finally:
        cliente.app.dependency_overrides.clear()


def test_contrato_extrair_ia_com_citacao_e_confirma(monkeypatch):
    """Caminho feliz: IA local extrai com citação, pausa e retoma (ADR-0006)."""
    fake = FakeExtrator()
    cliente.app.dependency_overrides[contexto_extracao] = lambda: (CFG_TESTE, fake)
    monkeypatch.setattr("sidecar.app.extrair_texto_pdf_bytes", lambda _b: DOC_CONTRATO)
    try:
        resp = cliente.post(
            "/contrato/extrair",
            json={"pdf_base64": PDF_B64, "nome": "contrato.pdf"},
            headers=CABECALHO,
        )
        assert resp.status_code == 200
        dados = resp.json()
        assert dados["modo"] == "ia"
        assert dados["thread_id"]
        assert dados["descartados"] == []

        campos = {c["chave"]: c for c in dados["campos"]}
        assert {"credor", "saldo", "taxa", "parcela", "restantes"} <= set(campos)
        # Valor no formato do formulário + citação verbatim do documento.
        assert campos["saldo"]["valor"] == "10000,00"
        assert "10.000,00" in campos["saldo"]["fonte"]
        assert campos["taxa"]["valor"] == "2,00"  # fração 0.02 → 2,00%

        # interrupt→resume: a confirmação retoma o grafo pausado.
        resp2 = cliente.post(
            "/contrato/confirmar",
            json={"thread_id": dados["thread_id"], "confirmacao": {"saldo": "10000,00"}},
            headers=CABECALHO,
        )
        assert resp2.status_code == 200
        assert resp2.json()["ok"] is True
        assert fake.chamadas == 1
    finally:
        cliente.app.dependency_overrides.clear()


def test_contrato_extrair_fallback_classico(monkeypatch):
    """IA local indisponível ⇒ extração clássica por regex, sem citação."""
    fake = FakeExtrator(erro=ValueError("sem llm local"))
    cliente.app.dependency_overrides[contexto_extracao] = lambda: (CFG_TESTE, fake)
    monkeypatch.setattr("sidecar.app.extrair_texto_pdf_bytes", lambda _b: DOC_CONTRATO)
    try:
        resp = cliente.post(
            "/contrato/extrair", json={"pdf_base64": PDF_B64}, headers=CABECALHO
        )
        assert resp.status_code == 200
        dados = resp.json()
        assert dados["modo"] == "classico"
        assert dados["thread_id"] is None
        chaves = {c["chave"] for c in dados["campos"]}
        # O regex acha taxa e parcelas no DOC; nenhum campo traz citação.
        assert {"taxa", "restantes"} <= chaves
        assert all(c["fonte"] == "" for c in dados["campos"])
        # Diagnóstico: o motivo da queda e o alvo efetivo da LLM (ADR-0010).
        assert "ERRO_PROVIDER:ValueError" in dados["motivos"]
        assert dados["llm"]["endpoint_local"] is True
        assert "1234" not in dados["llm"]["base_url"]  # CFG_TESTE usa o default
    finally:
        cliente.app.dependency_overrides.clear()


def test_fusao_classica_completa_campos_da_ia():
    """LLM pequena devolve null em campos presentes no doc ⇒ resgate sem LLM."""
    from sidecar.app import _fundir_com_classico

    campos = {
        "credor": {"valor": "Banco X", "trecho_fonte": "Credor: Banco X",
                   "confianca": 0.9},
        "tipo": None,
        "saldo_devedor": None,
        "taxa_mensal": {"valor": 0.0142, "trecho_fonte": "1,42% ao mês",
                        "confianca": 0.9},
        "parcela": {"valor": 899.47, "trecho_fonte": "96x de R$ 899,47",
                    "confianca": 0.9},
        "parcelas_restantes": None,
    }
    texto = ("Contrato de empréstimo consignado\n"
             "Total financiado Período de pagamento\n"
             "R$ 46.533,20 Agosto/2026 a Julho/2034\n")
    fundido = _fundir_com_classico(campos, texto)

    # `restantes` derivado do trecho já citado (e verificado) da parcela.
    assert fundido["parcelas_restantes"]["valor"] == 96
    assert fundido["parcelas_restantes"]["trecho_fonte"] == "96x de R$ 899,47"
    # `saldo` e `tipo` resgatados pelo regex clássico (sem citação).
    assert fundido["saldo_devedor"]["valor"] == 46533.20
    assert fundido["saldo_devedor"]["trecho_fonte"] == ""
    assert fundido["tipo"]["valor"] == "Consignado"
    # O que a IA achou nunca é sobrescrito.
    assert fundido["parcela"]["valor"] == 899.47
    assert fundido["credor"]["valor"] == "Banco X"


def test_contexto_trunca_para_openai_compat_local():
    """LM Studio (sem embeddings): trunca o documento em vez de tentar /api/embed."""
    from agent.config import ConfigAgente
    from sidecar.app import LIMITE_EXTRACAO_LLM, _contexto_seguro

    longo = "x" * (LIMITE_EXTRACAO_LLM + 500)
    cfg = ConfigAgente(provider="openai_compat", base_url="http://localhost:1234/v1")
    assert _contexto_seguro(longo, cfg) == longo[:LIMITE_EXTRACAO_LLM]


# --- Tela Análise: pacote determinístico (T-902, REQ-F-015) -------------------

PERFIL_ANALISE = {
    "renda": {"salario_liquido": 6000.0},
    "fixas": {"moradia": 1500.0},
    "dividas": [
        {
            "credor": "Cartão Banco João da Silva",
            "tipo": "Cartão de crédito",
            "saldo_devedor": 4000.0,
            "taxa_mensal": 0.12,
            "parcela": 800.0,
            "parcelas_restantes": 10,
        },
        {
            "credor": "Consignado Maria Pereira",
            "tipo": "Consignado",
            "saldo_devedor": 6000.0,
            "taxa_mensal": 0.015,
            "parcela": 350.0,
            "parcelas_restantes": 20,
        },
    ],
}


def test_analise_sem_token_401():
    resposta = cliente.post("/analise", json={"perfil": {}})
    assert resposta.status_code == 401


def test_analise_pacote_deterministico():
    payload = {"perfil": PERFIL_ANALISE, "extra": 300.0, "taxa_alvo": 0.018}
    resposta = cliente.post("/analise", json=payload, headers=CABECALHO)
    assert resposta.status_code == 200
    dados = resposta.json()

    # Estratégias recalculadas com o extra; avalanche nunca paga mais juros.
    ava = dados["estrategias"]["avalanche"]
    bola = dados["estrategias"]["bola_de_neve"]
    assert ava["quitavel"] and bola["quitavel"]
    assert dados["economia_avalanche"] == round(
        bola["juros_pagos"] - ava["juros_pagos"], 2)
    assert dados["economia_avalanche"] >= 0

    # Portabilidade: só o cartão (12% a.m.) supera a taxa-alvo de 1,8% a.m.
    ops = dados["oportunidades"]
    assert [o["credor"] for o in ops] == ["Cartão Banco João da Silva"]
    assert ops[0]["taxa_mensal"] == 0.12
    assert ops[0]["parcelas_restantes"] == 10
    assert ops[0]["economia_mensal"] > 0
    assert dados["economia_total_portabilidade"] == ops[0]["economia_total"]

    assert dados["recomendacoes"]  # regras do core sempre orientam algo


def test_analise_taxa_alvo_alta_sem_oportunidades():
    payload = {"perfil": PERFIL_ANALISE, "taxa_alvo": 0.20}
    resposta = cliente.post("/analise", json=payload, headers=CABECALHO)
    assert resposta.status_code == 200
    dados = resposta.json()
    assert dados["oportunidades"] == []
    assert dados["economia_total_portabilidade"] == 0.0


# --- Análise sênior: job assíncrono + fronteira cloud (H2/SEC-003) -------------


class ProviderEspiao:
    """Grava o payload que REALMENTE chegaria ao LLM (a fronteira cloud)."""

    def __init__(self, erro: Exception | None = None):
        self.payloads: list[str] = []
        self._fake = FakeProvider()
        self._erro = erro

    def analisar(self, fatos: FatosFinanceiros) -> AnaliseAgente:
        self.payloads.append(fatos.model_dump_json())
        if self._erro:
            raise self._erro
        return self._fake.analisar(fatos)


def _esperar_job(job_id: str, timeout_s: float = 5.0) -> dict:
    """Faz poll até o job sair de `rodando` (o teste roda com FakeProvider)."""
    fim = time.monotonic() + timeout_s
    while time.monotonic() < fim:
        resp = cliente.get(f"/analise/ia/{job_id}", headers=CABECALHO)
        assert resp.status_code == 200
        dados = resp.json()
        if dados["status"] != "rodando":
            return dados
        time.sleep(0.05)
    raise AssertionError("job da IA não terminou no tempo do teste")


def test_analise_ia_job_completo_e_anonimizacao_da_fronteira():
    """H2/SEC-003: nomes reais NUNCA cruzam a fronteira do provider; a
    desanonimização acontece só na seção devolvida à tela (exibição local)."""
    espiao = ProviderEspiao()
    cfg = ConfigAgente(provider="fake", cache=False)
    cliente.app.dependency_overrides[contexto_analise] = lambda: (cfg, espiao)
    try:
        resp = cliente.post("/analise/ia",
                            json={"perfil": PERFIL_ANALISE, "extra": 300.0},
                            headers=CABECALHO)
        assert resp.status_code == 200
        dados = _esperar_job(resp.json()["job_id"])

        # O payload enviado ao LLM só tem tokens — nenhum credor real, nenhum CPF.
        assert len(espiao.payloads) == 1
        payload_llm = espiao.payloads[0]
        assert "CREDOR_1" in payload_llm and "CREDOR_2" in payload_llm
        assert "João" not in payload_llm
        assert "Maria" not in payload_llm

        # A seção exibida ao usuário volta com os nomes reais restaurados.
        assert dados["status"] == "pronto"
        secao = dados["secao"]
        assert secao["modo"] == "completo"
        texto_secao = str(secao)
        assert "Cartão Banco João da Silva" in texto_secao
        assert "CREDOR_1" not in texto_secao
        assert secao["aviso_legal"]
    finally:
        cliente.app.dependency_overrides.clear()


def test_analise_ia_provider_falho_degrada_sem_500():
    """P8: falha do LLM degrada para o determinístico — o job nunca vira 500."""
    espiao = ProviderEspiao(erro=ValueError("llm fora do ar"))
    cfg = ConfigAgente(provider="fake", cache=False)
    cliente.app.dependency_overrides[contexto_analise] = lambda: (cfg, espiao)
    try:
        resp = cliente.post("/analise/ia", json={"perfil": PERFIL_ANALISE},
                            headers=CABECALHO)
        dados = _esperar_job(resp.json()["job_id"])
        assert dados["status"] == "pronto"
        assert dados["secao"]["modo"] == "degradado"
        assert dados["secao"]["motivos"]  # diz por que degradou
    finally:
        cliente.app.dependency_overrides.clear()


def test_analise_ia_job_desconhecido_404():
    resp = cliente.get("/analise/ia/nao-existe", headers=CABECALHO)
    assert resp.status_code == 404


def _esperar_terminal_em_memoria(job_id: str, timeout_s: float = 5.0) -> None:
    """Espera o job virar terminal olhando `_JOBS_IA` DIRETO — sem passar pelo
    endpoint de status, que apagaria o job na leitura final (precisamos do job
    terminal ainda preso para exercitar o TTL)."""
    from sidecar import app as app_mod

    fim = time.monotonic() + timeout_s
    while time.monotonic() < fim:
        with app_mod._JOBS_IA_LOCK:
            job = app_mod._JOBS_IA.get(job_id)
            if job is not None and job["status"] != "rodando":
                return
        time.sleep(0.02)
    raise AssertionError("job da IA não virou terminal no tempo do teste")


def test_analise_ia_job_terminal_expira_por_ttl(monkeypatch):
    """C-04: um job terminal nunca lido no poll (GUI fechou / auto-lock / tela
    só polla o catálogo) é varrido no próximo acesso ao dicionário assim que
    passa o TTL — antes ficava preso em memória com a seção DESANONIMIZADA."""
    from sidecar import app as app_mod

    reloginho = {"t": 1_000.0}
    monkeypatch.setattr(app_mod, "_relogio_jobs", lambda: reloginho["t"])

    espiao = ProviderEspiao()
    cfg = ConfigAgente(provider="fake", cache=False)
    cliente.app.dependency_overrides[contexto_analise] = lambda: (cfg, espiao)
    try:
        job_id = cliente.post("/analise/ia", json={"perfil": PERFIL_ANALISE},
                              headers=CABECALHO).json()["job_id"]
        _esperar_terminal_em_memoria(job_id)
        with app_mod._JOBS_IA_LOCK:
            assert job_id in app_mod._JOBS_IA  # terminal, ainda não lido

        # Passa o TTL e faz outro acesso ao dicionário (novo job) — coleta.
        reloginho["t"] += app_mod._TTL_JOBS_S + 1.0
        cliente.post("/analise/ia", json={"perfil": PERFIL_ANALISE}, headers=CABECALHO)
        with app_mod._JOBS_IA_LOCK:
            assert job_id not in app_mod._JOBS_IA  # varrido (antes: continuava preso)
            assert job_id not in app_mod._JOBS_IA_FIM
    finally:
        cliente.app.dependency_overrides.clear()
        app_mod._descartar_jobs_ia()  # não deixa jobs deste teste vazarem


def test_descartar_jobs_ia_esvazia_dicionarios():
    """C-04: o gancho de bloqueio esvazia `_JOBS_IA` (PII desanonimizada não
    sobrevive à janela desbloqueada do cofre)."""
    from sidecar import app as app_mod

    with app_mod._JOBS_IA_LOCK:
        app_mod._JOBS_IA["j"] = {"status": "pronto",
                                 "secao": {"credor": "Maria Real"}, "erro": ""}
        app_mod._JOBS_IA_FIM["j"] = 0.0
    app_mod._descartar_jobs_ia()
    with app_mod._JOBS_IA_LOCK:
        assert app_mod._JOBS_IA == {}
        assert app_mod._JOBS_IA_FIM == {}


def test_job_ia_descartado_no_meio_nao_ressuscita_pii(monkeypatch):
    """C-04 (revisão): se o cofre bloqueia ENQUANTO o job roda, o worker não
    pode regravar a seção desanonimizada ao terminar — sem a guarda no
    `_rodar_job_ia`, a PII ressuscitava em `_JOBS_IA` depois do bloqueio."""
    from core.models import PerfilFinanceiro
    from sidecar import app as app_mod

    monkeypatch.setattr(
        app_mod, "analisar",
        lambda *a, **k: (_ for _ in ()).throw(RuntimeError("nem chega aqui")))
    with app_mod._JOBS_IA_LOCK:
        app_mod._JOBS_IA["j2"] = {"status": "rodando", "secao": None, "erro": ""}
    app_mod._descartar_jobs_ia()          # cofre bloqueou no meio do job
    app_mod._rodar_job_ia("j2", PerfilFinanceiro(), 0.0, None, None)
    with app_mod._JOBS_IA_LOCK:
        assert "j2" not in app_mod._JOBS_IA      # resultado morreu com o descarte
        assert "j2" not in app_mod._JOBS_IA_FIM


# --- Carta ao credor (T-903, REQ-F-016) ---------------------------------------

DIVIDA_CARTA = {
    "credor": "Banco Alfa",
    "tipo": "Cartão de crédito",
    "saldo_devedor": 4500.0,
    "taxa_mensal": 0.11,
    "parcela": 700.0,
    "parcelas_restantes": 9,
}


def test_carta_sem_token_401():
    resposta = cliente.post("/carta/previa", json={"divida": DIVIDA_CARTA})
    assert resposta.status_code == 401


def test_carta_previa_quitacao_cita_valor_proposto():
    resposta = cliente.post(
        "/carta/previa",
        json={"divida": DIVIDA_CARTA, "tipo": "quitacao",
              "valor_proposto": 3000.0, "nome_usuario": "Fulana de Tal",
              "contrato": "000123-4"},
        headers=CABECALHO,
    )
    assert resposta.status_code == 200
    carta = resposta.json()
    assert carta["titulo"] == "Proposta de quitação à vista"
    assert carta["destinatario"] == "Banco Alfa"
    assert "Contrato nº 000123-4" in carta["referencia"]
    assert carta["assinatura"] == "Fulana de Tal"
    texto = " ".join(carta["paragrafos"])
    assert "R$ 4.500,00" in texto      # saldo da dívida
    assert "R$ 3.000,00" in texto      # valor proposto à vista


def test_carta_previa_portabilidade_cita_banco_e_taxa():
    resposta = cliente.post(
        "/carta/previa",
        json={"divida": DIVIDA_CARTA, "tipo": "portabilidade",
              "banco_concorrente": "Banco Beta",
              "taxa_concorrente_mensal": 0.018},
        headers=CABECALHO,
    )
    assert resposta.status_code == 200
    texto = " ".join(resposta.json()["paragrafos"])
    assert "Banco Beta" in texto
    assert "1,8" in texto              # taxa formatada em % a.m.


def test_carta_previa_tipo_desconhecido_cai_em_quitacao():
    resposta = cliente.post(
        "/carta/previa",
        json={"divida": DIVIDA_CARTA, "tipo": "inexistente"},
        headers=CABECALHO,
    )
    assert resposta.status_code == 200
    assert resposta.json()["tipo"] == "quitacao"


def test_exportar_carta_docx(tmp_path):
    docx = tmp_path / "proposta.docx"
    resposta = cliente.post(
        "/exportar/carta",
        json={"divida": DIVIDA_CARTA, "tipo": "reducao",
              "caminho": str(docx), "nome_usuario": "Fulana"},
        headers=CABECALHO,
    )
    assert resposta.status_code == 200
    assert resposta.json()["caminho"] == str(docx)
    assert docx.stat().st_size > 0


# --- Exportações .xlsx/.docx (T-902) ------------------------------------------


def test_exportar_planilha_e_relatorio(tmp_path, repo_tmp):
    xlsx = tmp_path / "diagnostico.xlsx"
    resp = cliente.post(
        "/exportar/planilha",
        json={"perfil": PERFIL_ANALISE, "caminho": str(xlsx),
              "extra": 300.0, "taxa_alvo": 0.018},
        headers=CABECALHO,
    )
    assert resp.status_code == 200
    assert resp.json()["caminho"] == str(xlsx)
    assert xlsx.stat().st_size > 0

    docx = tmp_path / "relatorio.docx"
    secao_ia = {"modo": "completo", "sumario": "Resumo de teste.",
                "diagnostico": "Diagnóstico de teste.", "confianca": 0.8,
                "aviso_legal": "Conteúdo assistido por IA."}
    resp = cliente.post(
        "/exportar/relatorio",
        json={"perfil": PERFIL_ANALISE, "caminho": str(docx),
              "nome_usuario": "Fulano", "secao_ia": secao_ia},
        headers=CABECALHO,
    )
    assert resp.status_code == 200
    assert docx.stat().st_size > 0


def test_exportar_planilha_inclui_rubricas_salvas(tmp_path, repo_tmp):
    """As rubricas do banco entram na aba "Orçamento detalhado" (T-1105)."""
    from openpyxl import load_workbook

    _criar_rubrica("Conta de luz", 180.0)
    xlsx = tmp_path / "diagnostico.xlsx"
    resp = cliente.post(
        "/exportar/planilha",
        json={"perfil": PERFIL_ANALISE, "caminho": str(xlsx)},
        headers=CABECALHO,
    )
    assert resp.status_code == 200
    ws = load_workbook(xlsx)["Orçamento detalhado"]
    textos = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
    assert any("Conta de luz" in t for t in textos)


def test_exportar_planilha_inclui_historico_arquivado(tmp_path, repo_tmp):
    """As competências do banco entram na aba "Evolução mensal" (T-1305)."""
    from openpyxl import load_workbook

    cliente.post("/estado", json={"variaveis": {"mercado": 750.0}},
                 headers=CABECALHO)
    cliente.post("/historico/arquivar", json={"mes": "2026-05"},
                 headers=CABECALHO)
    xlsx = tmp_path / "diagnostico.xlsx"
    resp = cliente.post(
        "/exportar/planilha",
        json={"perfil": PERFIL_ANALISE, "caminho": str(xlsx)},
        headers=CABECALHO,
    )
    assert resp.status_code == 200
    ws = load_workbook(xlsx)["Evolução mensal"]
    textos = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
    assert "2026-05" in textos
    assert any("Mercado" in t for t in textos)


def test_exportar_caminho_invalido_400(repo_tmp):
    resp = cliente.post(
        "/exportar/planilha",
        json={"perfil": PERFIL_ANALISE,
              "caminho": "Z:/pasta/que/nao/existe/x.xlsx"},
        headers=CABECALHO,
    )
    assert resp.status_code == 400
    assert "salvar" in resp.json()["detail"]


def test_contrato_pdf_escaneado_sem_ocr_degrada(monkeypatch):
    """PDF escaneado (sem texto) + motor de OCR indisponível ⇒ aviso (P8)."""
    monkeypatch.setattr("sidecar.app.extrair_texto_pdf_bytes", lambda _b: "   ")
    monkeypatch.setattr("sidecar.app._motor_ocr_singleton", lambda: None)
    resp = cliente.post(
        "/contrato/extrair", json={"pdf_base64": PDF_B64}, headers=CABECALHO
    )
    assert resp.status_code == 200
    dados = resp.json()
    assert dados["modo"] == "vazio"
    assert dados["ocr"] is False
    assert "OCR_INDISPONIVEL" in dados["motivos"]
    assert dados["campos"] == []


def test_contrato_imagem_ocr_extrai(monkeypatch):
    """Imagem (JPG) ⇒ OCR local produz o texto, que segue para a extração (REQ-F-024)."""
    fake = FakeExtrator()
    # O motor de OCR falso devolve o texto do contrato (ignora os bytes da imagem).
    motor = MotorFalso([_linha(DOC_CONTRATO, topo=10, esquerda=10)])
    cliente.app.dependency_overrides[contexto_extracao] = lambda: (CFG_TESTE, fake)
    cliente.app.dependency_overrides[contexto_ocr] = lambda: motor
    try:
        resp = cliente.post(
            "/contrato/extrair",
            json={"pdf_base64": PDF_B64, "nome": "contrato.jpg"},
            headers=CABECALHO,
        )
        assert resp.status_code == 200
        dados = resp.json()
        assert dados["ocr"] is True
        assert dados["modo"] == "ia"
        assert motor.chamadas == 1
        # O texto do OCR realimenta a extração: os campos citados foram achados.
        campos = {c["chave"] for c in dados["campos"]}
        assert {"saldo", "taxa", "parcela"} <= campos
    finally:
        cliente.app.dependency_overrides.clear()


# --- Estado persistido (T-1102, REQ-F-018 / ADR-0012) -------------------------


@pytest.fixture()
def repo_tmp(tmp_path):
    """Banco isolado por teste — a dependência real nunca toca o do usuário.

    Sobrescreve `exigir_cofre` diretamente (não `sessao_dependencia`): os
    testes deste banco não querem saber de cofre/sessão, só do repositório —
    mesmo padrão de antes do T-1603, só que o alvo do override passou de
    `repositorio` para `exigir_cofre` (o gate 423 agora mora ali).
    """
    repo = Repositorio(tmp_path / "dados.db")
    app.dependency_overrides[exigir_cofre] = lambda: repo
    yield repo
    del app.dependency_overrides[exigir_cofre]
    repo.fechar()


def test_estado_sem_token_401(repo_tmp):
    assert cliente.get("/estado").status_code == 401
    assert cliente.post("/estado", json={}).status_code == 401


def test_estado_primeira_execucao_sem_perfil(repo_tmp):
    resposta = cliente.get("/estado", headers=CABECALHO)
    assert resposta.status_code == 200
    assert resposta.json()["perfil"] is None


def test_estado_roundtrip_salvar_e_hidratar(repo_tmp):
    payload = {
        "renda": {"salario_liquido": 4200.0},
        "fixas": {"contas_casa": 480.5},
        "reserva_emergencia": 1000.0,
        "dividas": [
            {"credor": "Banco São João", "tipo": "Cartão de crédito",
             "saldo_devedor": 12000.0, "taxa_mensal": 0.11,
             "parcela": 700.0, "parcelas_restantes": 10}
        ],
    }
    assert cliente.post("/estado", json=payload,
                        headers=CABECALHO).json() == {"ok": True}

    perfil = cliente.get("/estado", headers=CABECALHO).json()["perfil"]
    assert perfil["renda"]["salario_liquido"] == 4200.0
    assert perfil["reserva_emergencia"] == 1000.0
    assert perfil["dividas"][0]["credor"] == "Banco São João"
    # O payload é NORMALIZADO pelo PerfilIn antes de persistir: campos omitidos
    # voltam com os defaults do schema — a hidratação nunca surpreende a GUI.
    assert perfil["fixas"]["moradia"] == 0.0
    assert perfil["variaveis"]["mercado"] == 0.0


def test_estado_salvar_de_novo_sobrescreve(repo_tmp):
    cliente.post("/estado", json={"reserva_emergencia": 1.0}, headers=CABECALHO)
    cliente.post("/estado", json={"reserva_emergencia": 2.0}, headers=CABECALHO)
    perfil = cliente.get("/estado", headers=CABECALHO).json()["perfil"]
    assert perfil["reserva_emergencia"] == 2.0


def test_estado_payload_invalido_422(repo_tmp):
    # Dívida sem `credor`/`tipo` não passa da validação — e não vai ao banco.
    payload = {"dividas": [{"saldo_devedor": 100.0}]}
    resposta = cliente.post("/estado", json=payload, headers=CABECALHO)
    assert resposta.status_code == 422
    assert cliente.get("/estado", headers=CABECALHO).json()["perfil"] is None


# --- Rubricas (T-1103, REQ-F-017 / ADR-0012) ----------------------------------


def _criar_rubrica(nome: str, valor: float, campo="contas_casa",
                   categoria="fixas") -> dict:
    resposta = cliente.post(
        "/rubricas",
        json={"categoria": categoria, "campo_pai": campo,
              "nome": nome, "valor": valor},
        headers=CABECALHO,
    )
    assert resposta.status_code == 200
    return resposta.json()


def test_rubricas_sem_token_401(repo_tmp):
    assert cliente.get("/rubricas").status_code == 401
    assert cliente.post("/rubricas", json={}).status_code == 401


def test_rubrica_criar_aplica_roll_up_no_perfil(repo_tmp):
    # Perfil salvo com contas_casa digitado direto (500).
    cliente.post("/estado", json={"fixas": {"contas_casa": 500.0}},
                 headers=CABECALHO)

    dados = _criar_rubrica("Conta de luz", 180.0)
    assert dados["rubricas"][0]["id"] is not None
    # Campo detalhado passa a valer a SOMA das rubricas, não o valor digitado.
    assert dados["perfil"]["fixas"]["contas_casa"] == 180.0

    dados = _criar_rubrica("Internet", 120.0)
    assert dados["perfil"]["fixas"]["contas_casa"] == 300.0
    assert [r["nome"] for r in dados["rubricas"]] == ["Conta de luz", "Internet"]

    # O perfil PERSISTIDO ficou consistente (fonte dos demais endpoints).
    perfil = cliente.get("/estado", headers=CABECALHO).json()["perfil"]
    assert perfil["fixas"]["contas_casa"] == 300.0


def test_rubrica_editar_recalcula(repo_tmp):
    rid = _criar_rubrica("Luz", 180.0)["rubricas"][0]["id"]
    resposta = cliente.post(f"/rubricas/{rid}",
                            json={"nome": "Luz + taxa", "valor": 200.0},
                            headers=CABECALHO)
    assert resposta.status_code == 200
    dados = resposta.json()
    assert dados["rubricas"][0]["nome"] == "Luz + taxa"
    assert dados["perfil"]["fixas"]["contas_casa"] == 200.0


def test_rubrica_remover_mantem_a_ultima_soma(repo_tmp):
    rid = _criar_rubrica("Luz", 180.0)["rubricas"][0]["id"]
    resposta = cliente.post(f"/rubricas/{rid}/remover", headers=CABECALHO)
    assert resposta.status_code == 200
    dados = resposta.json()
    assert dados["rubricas"] == []
    # Sem rubricas o campo volta a ser editável, mas conserva a última soma —
    # remover o detalhamento não zera o orçamento (ADR-0012).
    assert dados["perfil"]["fixas"]["contas_casa"] == 180.0


def test_rubrica_mutacoes_devolvem_so_o_contrato(repo_tmp):
    """C-32 (baixo): as três rotas de mutação devolvem EXATAMENTE
    `{rubricas, perfil}` — sem as chaves `rubrica`/`ok`, que nenhuma tela
    consome (`RubricaMutOut` em `contract.ts` só modela essas duas). Antes da
    correção, `rubrica_criar`/`rubrica_editar` incluíam `rubrica` e
    `rubrica_remover` incluía `ok`."""
    criado = _criar_rubrica("Água", 90.0)
    assert set(criado.keys()) == {"rubricas", "perfil"}

    rid = criado["rubricas"][0]["id"]
    editado = cliente.post(f"/rubricas/{rid}",
                           json={"nome": "Água + esgoto", "valor": 95.0},
                           headers=CABECALHO).json()
    assert set(editado.keys()) == {"rubricas", "perfil"}

    removido = cliente.post(f"/rubricas/{rid}/remover", headers=CABECALHO).json()
    assert set(removido.keys()) == {"rubricas", "perfil"}


def test_rubrica_ancoragem_invalida_422(repo_tmp):
    r = cliente.post("/rubricas",
                     json={"categoria": "investimentos", "campo_pai": "acoes",
                           "nome": "PETR4"},
                     headers=CABECALHO)
    assert r.status_code == 422
    r = cliente.post("/rubricas",
                     json={"categoria": "fixas", "campo_pai": "mercado",
                           "nome": "Feira"},
                     headers=CABECALHO)
    assert r.status_code == 422
    assert "mercado" in r.json()["detail"]


def test_rubrica_id_desconhecido_404(repo_tmp):
    assert cliente.post("/rubricas/999", json={"nome": "X", "valor": 1.0},
                        headers=CABECALHO).status_code == 404
    assert cliente.post("/rubricas/999/remover",
                        headers=CABECALHO).status_code == 404


def test_estado_salvar_reimpoe_a_soma_das_rubricas(repo_tmp):
    _criar_rubrica("Luz", 180.0)
    # Um front fora de sincronia tenta gravar contas_casa=999 direto...
    cliente.post("/estado",
                 json={"fixas": {"contas_casa": 999.0, "moradia": 1400.0}},
                 headers=CABECALHO)
    perfil = cliente.get("/estado", headers=CABECALHO).json()["perfil"]
    # ...e o invariante vence: campo detalhado = soma; o resto fica como veio.
    assert perfil["fixas"]["contas_casa"] == 180.0
    assert perfil["fixas"]["moradia"] == 1400.0


def test_estado_devolve_as_rubricas_para_a_hidratacao(repo_tmp):
    _criar_rubrica("Luz", 180.0)
    dados = cliente.get("/estado", headers=CABECALHO).json()
    assert [r["nome"] for r in dados["rubricas"]] == ["Luz"]


# --- Histórico mensal (T-1202, REQ-F-019 / ADR-0013) --------------------------


def test_historico_arquivar_e_listar(repo_tmp):
    cliente.post("/estado", json={"variaveis": {"mercado": 800.0}},
                 headers=CABECALHO)
    _criar_rubrica("Luz", 180.0)

    resp = cliente.post("/historico/arquivar", json={"mes": "2026-06"},
                        headers=CABECALHO)
    assert resp.status_code == 200
    assert resp.json()["meses"] == ["2026-06"]
    assert cliente.get("/historico", headers=CABECALHO).json()["meses"] == [
        "2026-06"]

    snap = cliente.get("/historico/2026-06", headers=CABECALHO).json()
    # O snapshot leva o perfil INTEIRO (campos diretos + detalhados) e as
    # rubricas da competência.
    assert snap["perfil"]["variaveis"]["mercado"] == 800.0
    assert snap["perfil"]["fixas"]["contas_casa"] == 180.0  # roll-up do vivo
    assert [r["nome"] for r in snap["rubricas"]] == ["Luz"]


def test_historico_comparar_contra_o_vivo(repo_tmp):
    cliente.post("/estado", json={"variaveis": {"mercado": 800.0}},
                 headers=CABECALHO)
    cliente.post("/historico/arquivar", json={"mes": "2026-06"},
                 headers=CABECALHO)
    # O mês virou e o mercado subiu.
    cliente.post("/estado", json={"variaveis": {"mercado": 900.0}},
                 headers=CABECALHO)

    resp = cliente.post("/historico/comparar",
                        json={"mes_a": "2026-06", "mes_b": None},
                        headers=CABECALHO)
    assert resp.status_code == 200
    comp = resp.json()["comparacao"]
    variaveis = next(s for s in comp["secoes"] if s["categoria"] == "variaveis")
    mercado = next(c for c in variaveis["campos"] if c["campo"] == "mercado")
    assert mercado["delta"] == 100.0
    assert mercado["variacao_pct"] == 0.125  # "seu mercado subiu 12,5%"


def test_historico_comparar_entre_dois_meses(repo_tmp):
    cliente.post("/estado", json={"fixas": {"moradia": 1400.0}},
                 headers=CABECALHO)
    cliente.post("/historico/arquivar", json={"mes": "2026-05"},
                 headers=CABECALHO)
    cliente.post("/estado", json={"fixas": {"moradia": 1500.0}},
                 headers=CABECALHO)
    cliente.post("/historico/arquivar", json={"mes": "2026-06"},
                 headers=CABECALHO)

    resp = cliente.post("/historico/comparar",
                        json={"mes_a": "2026-05", "mes_b": "2026-06"},
                        headers=CABECALHO)
    fixas = next(s for s in resp.json()["comparacao"]["secoes"]
                 if s["categoria"] == "fixas")
    assert fixas["delta"] == 100.0


def test_historico_mes_invalido_422_e_sem_snapshot_404(repo_tmp):
    assert cliente.post("/historico/arquivar", json={"mes": "julho"},
                        headers=CABECALHO).status_code == 422
    assert cliente.get("/historico/2026-13",
                       headers=CABECALHO).status_code == 422
    assert cliente.get("/historico/2026-01",
                       headers=CABECALHO).status_code == 404
    assert cliente.post("/historico/comparar",
                        json={"mes_a": "2026-01"},
                        headers=CABECALHO).status_code == 404


def test_historico_evolucao_series_do_core(repo_tmp):
    # Duas competências: mercado 750 → 900.
    cliente.post("/estado", json={"variaveis": {"mercado": 750.0}},
                 headers=CABECALHO)
    cliente.post("/historico/arquivar", json={"mes": "2026-05"},
                 headers=CABECALHO)
    cliente.post("/estado", json={"variaveis": {"mercado": 900.0}},
                 headers=CABECALHO)
    cliente.post("/historico/arquivar", json={"mes": "2026-06"},
                 headers=CABECALHO)

    resp = cliente.get("/historico/evolucao", headers=CABECALHO)
    assert resp.status_code == 200
    serie = resp.json()
    assert serie["meses"] == ["2026-05", "2026-06"]
    variaveis = next(s for s in serie["secoes"]
                     if s["categoria"] == "variaveis")
    assert variaveis["totais"] == [750.0, 900.0]
    mercado = next(c for c in variaveis["campos"] if c["campo"] == "mercado")
    assert mercado["valores"] == [750.0, 900.0]


def test_historico_evolucao_sem_competencias(repo_tmp):
    # A rota literal vence o parâmetro {mes} (nunca 422 "competência inválida").
    resp = cliente.get("/historico/evolucao", headers=CABECALHO)
    assert resp.status_code == 200
    dados = resp.json()
    assert dados["meses"] == []
    assert [s["categoria"] for s in dados["secoes"]] == [
        "renda", "fixas", "variaveis"]


# --- Importação de CSV (T-1302, REQ-F-021 / ADR-0014) --------------------------

# Grupos resultantes (ordenados por total decrescente pelo core):
# 0 = Salário Acme Ltda (crédito, 3500) · 1 = Conta de luz Enel (débito,
# 180,50) · 2 = Uber Trip (débito, 137,40 em 2 lançamentos).
CSV_EXTRATO = (
    "Data,Descrição,Valor\n"
    "05/06/2026,Conta de luz Enel,-180.50\n"
    "07/06/2026,UBER *TRIP 8291,-92.30\n"
    "15/06/2026,UBER *TRIP 4415,-45.10\n"
    "30/06/2026,Salário Acme Ltda,3500.00\n"
)
CSV_B64 = base64.b64encode(CSV_EXTRATO.encode("utf-8")).decode()


def _importar_csv(fake, csv_b64=CSV_B64):
    cliente.app.dependency_overrides[contexto_classificacao] = (
        lambda: (CFG_TESTE, fake))
    try:
        return cliente.post("/importar/csv", json={"csv_base64": csv_b64},
                            headers=CABECALHO)
    finally:
        del cliente.app.dependency_overrides[contexto_classificacao]


def test_importar_csv_classifica_para_revisao(repo_tmp):
    fake = FakeClassificador(ClassificacaoExtrato(itens=[
        ItemClassificado(indice=0, categoria="renda",
                         campo_pai="salario_liquido"),
        ItemClassificado(indice=1, categoria="fixas", campo_pai="contas_casa"),
        ItemClassificado(indice=2, categoria="fixas", campo_pai="transporte"),
    ]))
    resp = _importar_csv(fake)
    assert resp.status_code == 200
    dados = resp.json()
    assert dados["modo"] == "ia"
    assert dados["competencia_sugerida"] == "2026-06"
    # Todo número vem do parser determinístico — a LLM só pôs os rótulos.
    salario, luz, uber = dados["grupos"]
    assert (salario["nome"], salario["total"], salario["quantidade"]) == (
        "Salário Acme Ltda", 3500.0, 1)
    assert salario["natureza"] == "credito"
    assert (salario["categoria"], salario["campo_pai"]) == (
        "renda", "salario_liquido")
    assert (uber["nome"], uber["total"], uber["quantidade"]) == (
        "Uber Trip", 137.4, 2)
    assert (uber["categoria"], uber["campo_pai"]) == ("fixas", "transporte")
    assert luz["campo_pai"] == "contas_casa"
    # E NADA foi persistido: a revisão humana vem antes de aplicar.
    assert cliente.get("/rubricas", headers=CABECALHO).json()["rubricas"] == []


def test_importar_csv_sem_llm_degrada_para_manual(repo_tmp):
    fake = FakeClassificador(erro=ValueError("porta fechada"))
    dados = _importar_csv(fake).json()
    assert dados["modo"] == "manual"
    assert dados["motivos"] == ["ERRO_PROVIDER:ValueError"]
    # Os grupos chegam do mesmo jeito (parse determinístico) — sem rótulo.
    assert len(dados["grupos"]) == 3
    assert all(g["categoria"] is None for g in dados["grupos"])


def test_importar_csv_rotulo_invalido_e_descartado(repo_tmp):
    # Crédito (salário) classificado como despesa: a trava derruba o item.
    fake = FakeClassificador(ClassificacaoExtrato(itens=[
        ItemClassificado(indice=0, categoria="variaveis", campo_pai="lazer"),
        ItemClassificado(indice=2, categoria="variaveis", campo_pai="lazer"),
    ]))
    dados = _importar_csv(fake).json()
    assert dados["descartes"] == ["0:NATUREZA:credito"]
    assert dados["grupos"][0]["categoria"] is None
    assert dados["grupos"][2]["campo_pai"] == "lazer"


def test_importar_csv_base64_invalido_422(repo_tmp):
    fake = FakeClassificador()
    assert _importar_csv(fake, csv_b64="###").status_code == 422
    assert fake.chamadas == 0


# --- Importação de comprovante ESCANEADO via OCR (T-1405, REQ-F-026) ----------

# Bytes ignorados pelo motor falso; o nome é imagem ⇒ segue o caminho de OCR.
IMG_B64 = base64.b64encode(b"imagem-escaneada").decode()

# Texto que o OCR "leria" de um extrato escaneado: mesmas 3 rubricas do teste de
# CSV (luz, Uber ×2, salário) + cabeçalho e linha de SALDO que NÃO viram
# lançamento — prova que o pipeline pós-parse é o mesmo do CSV.
OCR_EXTRATO_TXT = (
    "BANCO EXEMPLO S.A.\n"
    "05/06/2026 Conta de luz Enel -180,50\n"
    "07/06/2026 UBER *TRIP 8291 -92,30\n"
    "15/06/2026 UBER *TRIP 4415 -45,10\n"
    "30/06/2026 Salario Acme Ltda 3.500,00\n"
    "Saldo final: 3.244,50\n"
)


def _importar_ocr(fake, motor, arquivo_b64=IMG_B64, nome="extrato.jpg"):
    # Deletes pontuais (não .clear()): o fixture repo_tmp instala o override do
    # `exigir_cofre` e faz `del` dele no teardown — um clear() aqui o derrubaria.
    cliente.app.dependency_overrides[contexto_classificacao] = (
        lambda: (CFG_TESTE, fake))
    cliente.app.dependency_overrides[contexto_ocr] = lambda: motor
    try:
        return cliente.post("/importar/ocr",
                            json={"arquivo_base64": arquivo_b64, "nome": nome},
                            headers=CABECALHO)
    finally:
        del cliente.app.dependency_overrides[contexto_classificacao]
        del cliente.app.dependency_overrides[contexto_ocr]


def test_importar_ocr_classifica_para_revisao(repo_tmp):
    fake = FakeClassificador(ClassificacaoExtrato(itens=[
        ItemClassificado(indice=0, categoria="renda",
                         campo_pai="salario_liquido"),
        ItemClassificado(indice=1, categoria="fixas", campo_pai="contas_casa"),
        ItemClassificado(indice=2, categoria="fixas", campo_pai="transporte"),
    ]))
    motor = MotorFalso([_linha(OCR_EXTRATO_TXT, topo=10, esquerda=10)])
    dados = _importar_ocr(fake, motor).json()
    assert dados["ocr"] is True
    assert dados["modo"] == "ia"
    assert dados["competencia_sugerida"] == "2026-06"
    assert motor.chamadas == 1
    # Mesmos grupos/valores do CSV (o parse é o mesmo core); a LLM só rotulou.
    salario, luz, uber = dados["grupos"]
    assert (salario["nome"], salario["total"], salario["natureza"]) == (
        "Salario Acme Ltda", 3500.0, "credito")
    assert (uber["nome"], uber["total"], uber["quantidade"]) == (
        "Uber Trip", 137.4, 2)
    assert uber["campo_pai"] == "transporte" and luz["campo_pai"] == "contas_casa"
    # Nada persistido — a revisão humana (e o /importar/aplicar) vêm depois.
    assert cliente.get("/rubricas", headers=CABECALHO).json()["rubricas"] == []


def test_importar_ocr_sem_motor_degrada(monkeypatch, repo_tmp):
    """Sem motor de OCR ⇒ degrada (P8): vazio, ocr=False, motivo — nunca 500."""
    # contexto_ocr default = None ⇒ cai no singleton; patchamos p/ indisponível
    # (mesmo padrão do teste de contrato escaneado sem OCR).
    monkeypatch.setattr("sidecar.app._motor_ocr_singleton", lambda: None)
    fake = FakeClassificador()
    cliente.app.dependency_overrides[contexto_classificacao] = (
        lambda: (CFG_TESTE, fake))
    try:
        resp = cliente.post("/importar/ocr",
                            json={"arquivo_base64": IMG_B64, "nome": "extrato.jpg"},
                            headers=CABECALHO)
    finally:
        del cliente.app.dependency_overrides[contexto_classificacao]
    dados = resp.json()
    assert dados["modo"] == "vazio"
    assert dados["ocr"] is False
    assert dados["motivos"] == ["OCR_INDISPONIVEL"]
    assert dados["grupos"] == []
    assert fake.chamadas == 0  # nem chega a classificar


def test_importar_ocr_base64_invalido_422(repo_tmp):
    fake = FakeClassificador()
    motor = MotorFalso([_linha("x", topo=1, esquerda=1)])
    assert _importar_ocr(fake, motor, arquivo_b64="###").status_code == 422


def test_importar_aplicar_no_vivo_faz_roll_up(repo_tmp):
    resp = cliente.post("/importar/aplicar", json={"mes": None, "itens": [
        {"categoria": "fixas", "campo_pai": "transporte",
         "nome": "Uber Trip", "valor": 137.40},
        {"categoria": "fixas", "campo_pai": "contas_casa",
         "nome": "Conta de luz Enel", "valor": 180.50},
    ]}, headers=CABECALHO)
    assert resp.status_code == 200
    dados = resp.json()
    assert dados["mes"] is None
    # Fluxo normal do ADR-0012: rubricas no vivo + roll-up na escrita.
    assert dados["perfil"]["fixas"]["transporte"] == 137.40
    assert dados["perfil"]["fixas"]["contas_casa"] == 180.50
    assert [r["nome"] for r in dados["rubricas"]] == [
        "Conta de luz Enel", "Uber Trip"]


def test_importar_aplicar_em_competencia_nova(repo_tmp):
    resp = cliente.post("/importar/aplicar", json={"mes": "2026-06", "itens": [
        {"categoria": "variaveis", "campo_pai": "mercado",
         "nome": "Mercado Bom Preço", "valor": 800.87},
    ]}, headers=CABECALHO)
    assert resp.status_code == 200
    assert resp.json()["meses"] == ["2026-06"]

    # O snapshot nasceu com o perfil recalculado (base zerada + soma).
    snap = cliente.get("/historico/2026-06", headers=CABECALHO).json()
    assert snap["perfil"]["variaveis"]["mercado"] == 800.87
    assert [r["nome"] for r in snap["rubricas"]] == ["Mercado Bom Preço"]
    # O orçamento VIVO não foi tocado.
    assert cliente.get("/rubricas", headers=CABECALHO).json()["rubricas"] == []
    assert cliente.get("/estado", headers=CABECALHO).json()["perfil"] is None


def test_importar_aplicar_acrescenta_sem_apagar(repo_tmp):
    # Competência já arquivada com uma rubrica viva copiada.
    cliente.post("/estado", json={"variaveis": {"mercado": 100.0}},
                 headers=CABECALHO)
    _criar_rubrica("Feira", 100.0, campo="mercado", categoria="variaveis")
    cliente.post("/historico/arquivar", json={"mes": "2026-06"},
                 headers=CABECALHO)

    # Importar o cartão do mesmo mês ACRESCENTA ao snapshot (ADR-0014).
    cliente.post("/importar/aplicar", json={"mes": "2026-06", "itens": [
        {"categoria": "variaveis", "campo_pai": "mercado",
         "nome": "Mercado Bom Preço", "valor": 800.87},
    ]}, headers=CABECALHO)
    snap = cliente.get("/historico/2026-06", headers=CABECALHO).json()
    assert [r["nome"] for r in snap["rubricas"]] == [
        "Feira", "Mercado Bom Preço"]
    assert snap["perfil"]["variaveis"]["mercado"] == 900.87


def test_importar_aplicar_validacoes_422(repo_tmp):
    assert cliente.post("/importar/aplicar", json={"itens": []},
                        headers=CABECALHO).status_code == 422
    assert cliente.post("/importar/aplicar", json={"itens": [
        {"categoria": "fixas", "campo_pai": "acoes", "nome": "X"},
    ]}, headers=CABECALHO).status_code == 422
    assert cliente.post("/importar/aplicar", json={"mes": "junho", "itens": [
        {"categoria": "fixas", "campo_pai": "moradia", "nome": "Aluguel"},
    ]}, headers=CABECALHO).status_code == 422
